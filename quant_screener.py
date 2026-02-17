# =========================================================
# quant_screener.py  (v6)
# ---------------------------------------------------------
# [업데이트 내용]
# v5 → v6 변경:
#   4. GARP 스크리닝 추가: PEG<1.5 + 매출성장 10%↑ + ROE 12%↑
#   5. 캐시카우 스크리닝 추가: 영업CF 우량 + 저부채 + 이익률 높음
#   6. 턴어라운드 스크리닝 추가: 적자→흑자 전환 or 이익률 급개선
#   7. analyze_one_stock에 턴어라운드/캐시카우 감지 필드 추가
#   8. calc_valuation에 PSR, FCF수익률, PEG 등 파생지표 추가
#
# 출력 파일 (6개):
#   1) quant_all_stocks.xlsx   — 전체 종목
#   2) quant_screened.xlsx     — 우량주/저평가 스크리닝
#   3) quant_momentum.xlsx     — 폭발적 성장+마진개선
#   4) quant_GARP.xlsx         — 성장+합리적 가격 (피터 린치)
#   5) quant_cashcow.xlsx      — 현금흐름 우량 (버핏)
#   6) quant_turnaround.xlsx   — 실적 반등 종목
# =========================================================

import sys
import logging
import re
from pathlib import Path

import numpy as np
import pandas as pd

import config

try:
    from tqdm import tqdm
except ImportError:
    def tqdm(iterable, desc="", **kwargs):
        items = list(iterable)
        total = len(items)
        for i, item in enumerate(items):
            if i % max(1, total // 10) == 0:
                print(f"  {desc}: {i}/{total} ({i*100//total}%)")
            yield item
        print(f"  {desc}: {total}/{total} (100%)")

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
log = logging.getLogger("SCREENER")

DATA_DIR = config.DATA_DIR


# ─────────────────────────────────────────────
# 계정 매핑 (exact match용)
# ─────────────────────────────────────────────
EXACT_ACCOUNTS = {
    "매출액": ["매출액", "영업수익", "이자수익", "보험료수익", "순영업수익"],
    "영업이익": ["영업이익"],
    "순이익": ["지배주주순이익", "당기순이익"],
    "자본": ["자본", "자본총계", "지배주주지분", "지배기업주주지분"],
    "부채": ["부채", "부채총계"],
    "배당금": ["주당배당금"],
    # 캐시카우/턴어라운드용 추가 계정
    "영업CF": ["영업활동현금흐름", "영업활동으로인한현금흐름"],
    "투자CF": ["투자활동현금흐름", "투자활동으로인한현금흐름"],
    "CAPEX": ["유형자산의취득", "유형자산취득"],
    # Piotroski F-Score용 BS/IS 계정
    "자산총계": ["자산총계", "자산"],
    "유동자산": ["유동자산"],
    "유동부채": ["유동부채"],
    "매출총이익": ["매출총이익"],
}

EXCLUDE_KEYWORDS = [
    "증가율", "(-1Y)", "(평균)", "률(", "비율", "배율", "(-1A", "(-1Q", "/ 수정평균"
]


# ═════════════════════════════════════════════
# 유틸리티
# ═════════════════════════════════════════════

def normalize_code(x):
    try:
        if pd.isna(x) or str(x).strip() == "":
            return np.nan
        s = str(x).strip()
        if '.' in s:
            s = s.split('.')[0]
        return s.zfill(6)
    except (ValueError, AttributeError, TypeError):
        return np.nan


def load_table(prefix: str) -> pd.DataFrame:
    import db as _db
    df = _db.load_latest(prefix)

    if df.empty:
        return df

    df.columns = df.columns.str.strip()

    if "종목코드" in df.columns:
        df["종목코드"] = df["종목코드"].apply(normalize_code)
        df = df.dropna(subset=["종목코드"])

    # 기준일 정규화: "2023-12-31 00:00:00" → "2023-12-31"
    if "기준일" in df.columns:
        df["기준일"] = df["기준일"].astype(str).str[:10]

    for col in ["값", "종가", "시가총액", "상장주식수"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df


# 하위 호환용 별칭
load_csv = load_table


def _should_exclude(account_name: str) -> bool:
    for kw in EXCLUDE_KEYWORDS:
        if kw in account_name:
            return True
    return False


def find_account_value(df, target_key, date_filter=None):
    if df.empty or "계정" not in df.columns:
        return {}

    targets = EXACT_ACCOUNTS.get(target_key, [target_key])
    work = df.copy()
    if date_filter is not None:
        work = work[work["기준일"].isin(date_filter)]

    mask = work["계정"].isin(targets)
    matched = work[mask]

    if matched.empty:
        def _startswith_any(name):
            name_str = str(name)
            for t in targets:
                if name_str.startswith(t) and not _should_exclude(name_str):
                    return True
            return False
        mask2 = work["계정"].apply(_startswith_any)
        matched = work[mask2]

    if matched.empty:
        return {}

    matched = matched.drop_duplicates(["종목코드", "기준일"], keep="first")

    result = {}
    for _, r in matched.iterrows():
        try:
            dt = str(r["기준일"])
            v = float(r["값"]) if pd.notna(r["값"]) else None
            if v is not None:
                result[dt] = v
        except (ValueError, TypeError, KeyError):
            # 값 변환 실패 시 무시
            pass
    return result


# ═════════════════════════════════════════════
# 데이터 전처리 & 단위 감지
# ═════════════════════════════════════════════

def preprocess_indicators(ind_df):
    if ind_df.empty:
        return ind_df
    ind_df = ind_df.drop_duplicates(["종목코드", "기준일", "계정", "지표구분"], keep="first")
    return ind_df


def detect_unit_multiplier(ind_df):
    sam = ind_df[ind_df["종목코드"] == "005930"]
    if sam.empty:
        return 100_000_000

    sam_y = sam[sam["지표구분"] == "RATIO_Y"]
    rev_dict = find_account_value(sam_y, "매출액")

    if not rev_dict:
        return 100_000_000

    annual_revs = {d: v for d, v in rev_dict.items() if d.endswith("12-31")}
    if not annual_revs:
        annual_revs = rev_dict
    latest_rev = max(annual_revs.values())

    if latest_rev > 1e14: return 1
    elif latest_rev > 1e8: return 1_000_000
    elif latest_rev > 1e5: return 100_000_000
    else: return 100_000_000


# ═════════════════════════════════════════════
# 성장/추세 분석 유틸
# ═════════════════════════════════════════════

def calc_cagr(series_dict, min_years=2):
    if len(series_dict) < min_years:
        return np.nan
    dates = sorted(series_dict.keys())
    v0, v1 = series_dict[dates[0]], series_dict[dates[-1]]
    if v0 <= 0 or v1 <= 0:
        return np.nan
    try:
        d0, d1 = pd.Timestamp(dates[0]), pd.Timestamp(dates[-1])
        years = (d1 - d0).days / 365.25
        if years < 0.5: return np.nan
        return ((v1 / v0) ** (1 / years) - 1) * 100
    except:
        return np.nan


def count_consecutive_growth(series_dict):
    if len(series_dict) < 2:
        return 0
    vals = [series_dict[d] for d in sorted(series_dict.keys())]
    count = 0
    for i in range(len(vals) - 1, 0, -1):
        if vals[i] > vals[i - 1] and vals[i - 1] > 0:
            count += 1
        else:
            break
    return count


def _quarter_key(date_str):
    """분기 날짜 → (월-일) 키. 예: '2024-09-30' → '09-30'"""
    return date_str[5:10]


def _prev_year_date(date_str):
    """분기 날짜의 전년 동기 날짜. 예: '2024-09-30' → '2023-09-30'"""
    try:
        y = int(date_str[:4])
        return f"{y - 1}{date_str[4:]}"
    except (ValueError, IndexError):
        return None


def calc_quarterly_yoy(q_data, key):
    """분기별 전년동기비(YoY) 성장률을 계산.

    Returns:
        dict: {
            'latest_yoy': float or NaN (최근 분기 YoY %),
            'consecutive_yoy_growth': int (연속 YoY 성장 분기 수),
            'latest_quarter': str (최근 분기 날짜),
            'yoy_series': dict (날짜 → YoY% 시계열),
        }
    """
    result = {
        "latest_yoy": np.nan,
        "consecutive_yoy_growth": 0,
        "latest_quarter": "",
        "yoy_series": {},
    }

    q_dates = sorted(q_data["기준일"].unique())
    if len(q_dates) < 5:  # 최소 5개 분기 (4분기 + 전년 1개)
        return result

    # 분기별 값 추출
    vals = find_account_value(q_data, key)
    if len(vals) < 5:
        return result

    # YoY 계산: 각 분기에 대해 전년 동분기와 비교
    yoy_series = {}
    for date_str in sorted(vals.keys()):
        prev_date = _prev_year_date(date_str)
        if prev_date and prev_date in vals:
            prev_val = vals[prev_date]
            curr_val = vals[date_str]
            if prev_val > 0:
                yoy_series[date_str] = ((curr_val / prev_val) - 1) * 100
            elif prev_val < 0 and curr_val > 0:
                # 전년 적자 → 올해 흑자: 특수 케이스
                yoy_series[date_str] = np.nan  # 흑전은 %로 표현 부적절

    if not yoy_series:
        return result

    result["yoy_series"] = yoy_series
    result["latest_quarter"] = max(yoy_series.keys())
    result["latest_yoy"] = yoy_series[result["latest_quarter"]]

    # 연속 YoY 성장 카운트 (최근부터 역순)
    sorted_dates = sorted(yoy_series.keys(), reverse=True)
    count = 0
    for d in sorted_dates:
        yoy_val = yoy_series[d]
        if pd.notna(yoy_val) and yoy_val > 0:
            count += 1
        else:
            break
    result["consecutive_yoy_growth"] = count

    return result


def calc_ttm_yoy(q_data, key):
    """TTM(최근 4분기 합) vs 전년 TTM(1년 전 4분기 합) 비교.

    Returns:
        dict: {
            'ttm_current': float or NaN,
            'ttm_prev': float or NaN,
            'ttm_yoy': float or NaN (YoY %),
        }
    """
    result = {"ttm_current": np.nan, "ttm_prev": np.nan, "ttm_yoy": np.nan}

    q_dates = sorted(q_data["기준일"].unique())
    if len(q_dates) < 8:  # 최근 4분기 + 전년 4분기
        return result

    vals = find_account_value(q_data, key)
    if len(vals) < 8:
        return result

    sorted_dates = sorted(vals.keys())
    last4 = sorted_dates[-4:]
    prev4 = sorted_dates[-8:-4]

    ttm_curr = sum(vals[d] for d in last4 if d in vals)
    ttm_prev = sum(vals[d] for d in prev4 if d in vals)

    if len([d for d in last4 if d in vals]) == 4:
        result["ttm_current"] = ttm_curr
    if len([d for d in prev4 if d in vals]) == 4:
        result["ttm_prev"] = ttm_prev

    if pd.notna(result["ttm_current"]) and pd.notna(result["ttm_prev"]) and result["ttm_prev"] > 0:
        result["ttm_yoy"] = ((result["ttm_current"] / result["ttm_prev"]) - 1) * 100

    return result


# ═════════════════════════════════════════════
# 종목별 펀더멘털 분석 (v6: 턴어라운드/캐시카우 필드 추가)
# ═════════════════════════════════════════════

def analyze_one_stock(ticker, ind_grp, fs_grp):
    result = {"종목코드": ticker}
    has_ind = not ind_grp.empty and "지표구분" in ind_grp.columns
    has_fs = not fs_grp.empty

    # ── TTM (Trailing 12 Months) ──
    ttm_rev, ttm_op, ttm_ni = np.nan, np.nan, np.nan
    ttm_source = "없음"

    if has_ind:
        y_data = ind_grp[ind_grp["지표구분"] == "RATIO_Y"]
        q_data = ind_grp[ind_grp["지표구분"] == "RATIO_Q"]

        y_dates = sorted(y_data["기준일"].unique())
        annual_dates = [d for d in y_dates if str(d).endswith("12-31")]
        q_dates = sorted(q_data["기준일"].unique())
        last4q = q_dates[-4:] if len(q_dates) >= 4 else []

        for label, key, setter in [("매출", "매출액", "ttm_rev"), ("영업이익", "영업이익", "ttm_op"), ("순이익", "순이익", "ttm_ni")]:
            val = np.nan
            if last4q:
                d = find_account_value(q_data[q_data["기준일"].isin(last4q)], key)
                if len(d) >= 4: val = sum(d.values())
            if pd.isna(val) and annual_dates:
                d = find_account_value(y_data, key, annual_dates)
                if d: val = d[max(d.keys())]

            if setter == "ttm_rev": ttm_rev = val
            elif setter == "ttm_op": ttm_op = val
            else: ttm_ni = val

        if pd.notna(ttm_rev): ttm_source = "있음"

    result.update({"TTM_매출": ttm_rev, "TTM_영업이익": ttm_op, "TTM_순이익": ttm_ni, "TTM_소스": ttm_source})

    # ── 계절성 통제: 분기별 YoY 성장률 & TTM YoY ──
    if has_ind:
        q_data_all = ind_grp[ind_grp["지표구분"] == "RATIO_Q"]

        for label, key in [("매출", "매출액"), ("영업이익", "영업이익"), ("순이익", "순이익")]:
            # 분기별 YoY (전년동기비)
            qyoy = calc_quarterly_yoy(q_data_all, key)
            result[f"Q_{label}_YoY(%)"] = qyoy["latest_yoy"]
            result[f"Q_{label}_연속YoY성장"] = qyoy["consecutive_yoy_growth"]

            # TTM YoY (최근4분기 합 vs 전년4분기 합)
            ttm_yoy = calc_ttm_yoy(q_data_all, key)
            result[f"TTM_{label}_YoY(%)"] = ttm_yoy["ttm_yoy"]

        # 최근 분기 날짜 (참조용)
        q_dates_sorted = sorted(q_data_all["기준일"].unique())
        result["최근분기"] = q_dates_sorted[-1] if q_dates_sorted else ""
    else:
        for label in ["매출", "영업이익", "순이익"]:
            result[f"Q_{label}_YoY(%)"] = np.nan
            result[f"Q_{label}_연속YoY성장"] = 0
            result[f"TTM_{label}_YoY(%)"] = np.nan
        result["최근분기"] = ""

    # ── 자본/부채 ──
    curr_equity, curr_debt = np.nan, np.nan
    if has_fs:
        last_dt = sorted(fs_grp["기준일"].unique())[-1]
        bs_last = fs_grp[fs_grp["기준일"] == last_dt]
        e = find_account_value(bs_last, "자본")
        d = find_account_value(bs_last, "부채")
        if e: curr_equity = list(e.values())[0]
        if d: curr_debt = list(d.values())[0]

    if pd.isna(curr_equity) and has_ind:
        y_data = ind_grp[ind_grp["지표구분"] == "RATIO_Y"]
        e = find_account_value(y_data, "자본")
        d = find_account_value(y_data, "부채")
        if e: curr_equity = e[max(e.keys())]
        if d: curr_debt = d[max(d.keys())]

    result.update({"자본": curr_equity, "부채": curr_debt})

    # ── [v7] Piotroski F-Score용 BS/IS 시계열 ──
    total_assets_series, debt_series, equity_series = {}, {}, {}
    current_assets_series, current_liab_series = {}, {}
    gross_profit_series = {}

    if has_fs:
        fs_y = fs_grp[fs_grp["주기"] == "y"]
        total_assets_series = find_account_value(fs_y, "자산총계")
        current_assets_series = find_account_value(fs_y, "유동자산")
        current_liab_series = find_account_value(fs_y, "유동부채")
        gross_profit_series = find_account_value(fs_y, "매출총이익")
        debt_series = find_account_value(fs_y, "부채")
        equity_series = find_account_value(fs_y, "자본")

    # indicators fallback (BS 데이터)
    if not total_assets_series and has_ind:
        y_data = ind_grp[ind_grp["지표구분"] == "RATIO_Y"]
        total_assets_series = find_account_value(y_data, "자산총계")
    if not debt_series and has_ind:
        y_data = ind_grp[ind_grp["지표구분"] == "RATIO_Y"]
        debt_series = find_account_value(y_data, "부채")
        equity_series = find_account_value(y_data, "자본")

    result["자산총계"] = total_assets_series[max(total_assets_series.keys())] if total_assets_series else np.nan

    # ── 성장성 (CAGR) ──
    rev_series, op_series, ni_series = {}, {}, {}
    if has_ind:
        y_data = ind_grp[ind_grp["지표구분"] == "RATIO_Y"]
        annual_dates = [d for d in sorted(y_data["기준일"].unique()) if str(d).endswith("12-31")]
        if len(annual_dates) >= 2:
            rev_series = find_account_value(y_data, "매출액", annual_dates)
            op_series = find_account_value(y_data, "영업이익", annual_dates)
            ni_series = find_account_value(y_data, "순이익", annual_dates)

    result["매출_CAGR"] = calc_cagr(rev_series)
    result["영업이익_CAGR"] = calc_cagr(op_series)
    result["순이익_CAGR"] = calc_cagr(ni_series)
    result["매출_연속성장"] = count_consecutive_growth(rev_series)
    result["영업이익_연속성장"] = count_consecutive_growth(op_series)
    result["순이익_연속성장"] = count_consecutive_growth(ni_series)
    result["데이터_연수"] = len(rev_series)

    # ── 이익률 개선 여부 ──
    if len(rev_series) >= 2 and len(op_series) >= 2:
        latest = sorted(rev_series.keys())[-1]
        prev = sorted(rev_series.keys())[-2]
        opm_l = (op_series.get(latest, 0) / rev_series[latest] * 100) if rev_series[latest] > 0 else np.nan
        opm_p = (op_series.get(prev, 0) / rev_series[prev] * 100) if rev_series[prev] > 0 else np.nan
        result["영업이익률_최근"] = opm_l
        result["영업이익률_전년"] = opm_p
        result["이익률_개선"] = 1 if pd.notna(opm_l) and pd.notna(opm_p) and opm_l > opm_p else 0

        # [v6] 이익률 급개선 감지 (영업이익률 +5%p 이상)
        if pd.notna(opm_l) and pd.notna(opm_p):
            delta = opm_l - opm_p
            result["이익률_변동폭"] = delta
            result["이익률_급개선"] = 1 if delta >= 5 else 0
        else:
            result["이익률_변동폭"] = np.nan
            result["이익률_급개선"] = 0
    else:
        result.update({
            "영업이익률_최근": np.nan, "영업이익률_전년": np.nan,
            "이익률_개선": 0, "이익률_변동폭": np.nan, "이익률_급개선": 0,
        })

    # ── [v6] 턴어라운드 감지: 전년 순이익 < 0 → 올해 > 0 ──
    if len(ni_series) >= 2:
        ni_vals = [ni_series[d] for d in sorted(ni_series.keys())]
        result["순이익_전년음수"] = 1 if ni_vals[-2] < 0 else 0
        result["순이익_당기양수"] = 1 if ni_vals[-1] > 0 else 0
        result["흑자전환"] = 1 if ni_vals[-2] < 0 and ni_vals[-1] > 0 else 0
    else:
        result["순이익_전년음수"] = 0
        result["순이익_당기양수"] = 0
        result["흑자전환"] = 0

    # ── [v7] 영업CF / CAPEX / FCF 시계열 + TTM ──
    ocf_series, capex_series = {}, {}

    # 1) indicators(RATIO_Y)에서 연도별 시계열 추출
    if has_ind:
        y_data = ind_grp[ind_grp["지표구분"] == "RATIO_Y"]
        # annual_dates 명시적 정의 (12-31로 끝나는 기준일만)
        annual_dates = [d for d in sorted(y_data["기준일"].unique()) if str(d).endswith("12-31")]
        ocf_series = find_account_value(y_data, "영업CF", annual_dates)
        capex_series = find_account_value(y_data, "CAPEX", annual_dates)

    # 2) indicators에 없으면 financial_statements(CF)에서 fallback
    if not ocf_series and has_fs:
        fs_y = fs_grp[(fs_grp["주기"] == "y")]
        ocf_series = find_account_value(fs_y, "영업CF")
    if not capex_series and has_fs:
        fs_y = fs_grp[(fs_grp["주기"] == "y")]
        capex_series = find_account_value(fs_y, "CAPEX")

    # CAPEX는 FnGuide에서 음수로 기재되므로 절대값 처리
    capex_series = {d: abs(v) for d, v in capex_series.items()}

    # 3) FCF 시계열 (영업CF - CAPEX, 동일 연도만)
    fcf_series = {}
    common_dates = set(ocf_series.keys()) & set(capex_series.keys())
    for d in common_dates:
        fcf_series[d] = ocf_series[d] - capex_series[d]

    # TTM 값 (최신 연도)
    ttm_ocf = ocf_series[max(ocf_series.keys())] if ocf_series else np.nan
    ttm_capex = capex_series[max(capex_series.keys())] if capex_series else np.nan
    ttm_fcf = fcf_series[max(fcf_series.keys())] if fcf_series else np.nan

    result["TTM_영업CF"] = ttm_ocf
    result["TTM_CAPEX"] = ttm_capex
    result["TTM_FCF"] = ttm_fcf
    result["영업CF_CAGR"] = calc_cagr(ocf_series)
    result["FCF_CAGR"] = calc_cagr(fcf_series)
    result["영업CF_연속성장"] = count_consecutive_growth(ocf_series)

    # ── [v7] Piotroski F-Score (9개 항목, 각 0 or 1) ──
    f_score = 0

    # F1: ROA > 0 (순이익 > 0이면 자산 대비 수익 양)
    f1 = 1 if pd.notna(ttm_ni) and ttm_ni > 0 else 0

    # F2: 영업CF > 0
    f2 = 1 if pd.notna(ttm_ocf) and ttm_ocf > 0 else 0

    # F3: ROA 개선 (순이익/자산총계 전년 대비 개선)
    f3 = 0
    if len(ni_series) >= 2 and len(total_assets_series) >= 2:
        ni_dates = sorted(ni_series.keys())
        ta_dates = sorted(total_assets_series.keys())
        # 최근 2개년 공통
        if len(ni_dates) >= 2 and len(ta_dates) >= 2:
            roa_curr = ni_series[ni_dates[-1]] / total_assets_series[ta_dates[-1]] if total_assets_series[ta_dates[-1]] > 0 else 0
            roa_prev = ni_series[ni_dates[-2]] / total_assets_series[ta_dates[-2]] if total_assets_series[ta_dates[-2]] > 0 else 0
            f3 = 1 if roa_curr > roa_prev else 0

    # F4: 이익품질 (영업CF > 순이익)
    f4 = 1 if pd.notna(ttm_ocf) and pd.notna(ttm_ni) and ttm_ni > 0 and ttm_ocf > ttm_ni else 0

    # F5: 레버리지 감소 (부채비율 전년 대비 감소)
    f5 = 0
    if len(debt_series) >= 2 and len(equity_series) >= 2:
        d_dates = sorted(debt_series.keys())
        e_dates = sorted(equity_series.keys())
        if equity_series.get(e_dates[-1], 0) > 0 and equity_series.get(e_dates[-2], 0) > 0:
            dr_curr = debt_series[d_dates[-1]] / equity_series[e_dates[-1]]
            dr_prev = debt_series[d_dates[-2]] / equity_series[e_dates[-2]]
            f5 = 1 if dr_curr < dr_prev else 0

    # F6: 유동비율 개선 (유동자산/유동부채 전년 대비 개선)
    f6 = 0
    if len(current_assets_series) >= 2 and len(current_liab_series) >= 2:
        ca_dates = sorted(current_assets_series.keys())
        cl_dates = sorted(current_liab_series.keys())
        if current_liab_series.get(cl_dates[-1], 0) > 0 and current_liab_series.get(cl_dates[-2], 0) > 0:
            cr_curr = current_assets_series[ca_dates[-1]] / current_liab_series[cl_dates[-1]]
            cr_prev = current_assets_series[ca_dates[-2]] / current_liab_series[cl_dates[-2]]
            f6 = 1 if cr_curr > cr_prev else 0

    # F7: 주식 희석 없음 (발행주식수 미증가) — calc_valuation에서 처리 (shares 데이터 필요)
    f7 = 0  # placeholder, calc_valuation()에서 업데이트

    # F8: 매출총이익률 개선
    f8 = 0
    if len(gross_profit_series) >= 2 and len(rev_series) >= 2:
        gp_dates = sorted(gross_profit_series.keys())
        rv_dates = sorted(rev_series.keys())
        if len(gp_dates) >= 2 and len(rv_dates) >= 2:
            if rev_series.get(rv_dates[-1], 0) > 0 and rev_series.get(rv_dates[-2], 0) > 0:
                gm_curr = gross_profit_series[gp_dates[-1]] / rev_series[rv_dates[-1]]
                gm_prev = gross_profit_series[gp_dates[-2]] / rev_series[rv_dates[-2]]
                f8 = 1 if gm_curr > gm_prev else 0

    # F9: 자산회전율 개선 (매출/자산총계 전년 대비 개선)
    f9 = 0
    if len(rev_series) >= 2 and len(total_assets_series) >= 2:
        rv_dates = sorted(rev_series.keys())
        ta_dates = sorted(total_assets_series.keys())
        if total_assets_series.get(ta_dates[-1], 0) > 0 and total_assets_series.get(ta_dates[-2], 0) > 0:
            at_curr = rev_series[rv_dates[-1]] / total_assets_series[ta_dates[-1]]
            at_prev = rev_series[rv_dates[-2]] / total_assets_series[ta_dates[-2]]
            f9 = 1 if at_curr > at_prev else 0

    f_score = f1 + f2 + f3 + f4 + f5 + f6 + f7 + f8 + f9
    result["F스코어"] = f_score
    result["F1_수익성"] = f1
    result["F2_영업CF"] = f2
    result["F3_ROA개선"] = f3
    result["F4_이익품질"] = f4
    result["F5_레버리지"] = f5
    result["F6_유동성"] = f6
    result["F7_희석없음"] = f7
    result["F8_매출총이익률"] = f8
    result["F9_자산회전율"] = f9

    # ── 배당 ──
    dps_series = {}
    if has_ind:
        dps_data = ind_grp[ind_grp["지표구분"] == "DPS"]
        annual_dps = [d for d in sorted(dps_data["기준일"].unique()) if str(d).endswith("12-31")]
        if annual_dps:
            dps_series = find_account_value(dps_data, "배당금", annual_dps)

    result["DPS_최근"] = list(dps_series.values())[-1] if dps_series else np.nan
    result["DPS_CAGR"] = calc_cagr(dps_series)
    result["배당_연속증가"] = count_consecutive_growth(dps_series)

    # ── [v8] 배당 성장 추가 지표: 수익 + 배당 동반성장 ──
    # 순이익과 배당금이 함께 연속 증가하는지 체크
    result["배당_수익동반증가"] = 1 if (
        count_consecutive_growth(ni_series) >= 2 and
        count_consecutive_growth(dps_series) >= 1
    ) else 0

    return result


def analyze_all(fs_df, ind_df):
    results = []
    tickers = list(set(
        list(fs_df["종목코드"].unique() if not fs_df.empty else []) +
        list(ind_df["종목코드"].unique() if not ind_df.empty else [])
    ))
    for ticker in tqdm(tickers, desc="펀더멘털 분석", ncols=100):
        ind_grp = ind_df[ind_df["종목코드"] == ticker] if not ind_df.empty else pd.DataFrame()
        fs_grp = fs_df[fs_df["종목코드"] == ticker] if not fs_df.empty else pd.DataFrame()
        results.append(analyze_one_stock(ticker, ind_grp, fs_grp))
    return pd.DataFrame(results)


# ═════════════════════════════════════════════
# 밸류에이션 & 스코어링 (v6: 파생지표 추가)
# ═════════════════════════════════════════════

def calc_valuation(daily, anal_df, multiplier, shares_df):
    merge_cols = ["종목코드", "종목명", "종가", "시가총액", "상장주식수"]
    valid_merge = [c for c in merge_cols if c in daily.columns]
    df = daily[valid_merge].drop_duplicates("종목코드").merge(anal_df, on="종목코드", how="inner")

    if not shares_df.empty and "상장주식수" in df.columns:
        shares_map = shares_df.drop_duplicates("종목코드").set_index("종목코드")["발행주식수"]
        mask_no = df["상장주식수"].isna() | (df["상장주식수"] == 0)
        df.loc[mask_no, "상장주식수"] = df.loc[mask_no, "종목코드"].map(shares_map)

    M = multiplier

    # ── 기본 지표 ──
    df["PER"] = np.where((df["TTM_순이익"] > 0) & (df["시가총액"] > 0), df["시가총액"] / (df["TTM_순이익"] * M), np.nan)
    df["PBR"] = np.where((df["자본"] > 0) & (df["시가총액"] > 0), df["시가총액"] / (df["자본"] * M), np.nan)
    df["ROE(%)"] = np.where((df["자본"] > 0) & pd.notna(df["TTM_순이익"]), (df["TTM_순이익"] / df["자본"]) * 100, np.nan)
    df["부채비율(%)"] = np.where((df["자본"] > 0), (df["부채"] / df["자본"]) * 100, np.nan)
    df["영업이익률(%)"] = df["영업이익률_최근"]

    shares_safe = df["상장주식수"].replace(0, np.nan)
    df["BPS"] = (df["자본"] * M) / shares_safe
    df["EPS"] = (df["TTM_순이익"] * M) / shares_safe
    df["배당수익률(%)"] = np.where((df["종가"] > 0) & (df["DPS_최근"] > 0), (df["DPS_최근"] / df["종가"]) * 100, 0)

    # ── [v6] 추가 파생 지표 ──
    # PSR (Price-to-Sales)
    df["PSR"] = np.where(
        (df["TTM_매출"] > 0) & (df["시가총액"] > 0),
        df["시가총액"] / (df["TTM_매출"] * M), np.nan
    )

    # PEG (PER / 순이익CAGR) — GARP용
    df["PEG"] = np.where(
        pd.notna(df["PER"]) & (df["PER"] > 0) &
        pd.notna(df["순이익_CAGR"]) & (df["순이익_CAGR"] > 0),
        df["PER"] / df["순이익_CAGR"], np.nan
    )

    # 이익수익률 (Earnings Yield = EPS / 종가)
    df["이익수익률(%)"] = np.where(
        (df["종가"] > 0) & pd.notna(df["EPS"]) & (df["EPS"] > 0),
        (df["EPS"] / df["종가"]) * 100, np.nan
    )

    # FCF 수익률 (진짜 FCF = 영업CF - CAPEX)
    df["FCF수익률(%)"] = np.where(
        pd.notna(df["TTM_FCF"]) & (df["시가총액"] > 0),
        (df["TTM_FCF"] * M / df["시가총액"]) * 100, np.nan
    )

    # 현금전환율 (영업CF / 순이익 × 100, 100% 이상이면 이익이 현금으로 뒷받침됨)
    df["현금전환율(%)"] = np.where(
        pd.notna(df["TTM_영업CF"]) & pd.notna(df["TTM_순이익"]) & (df["TTM_순이익"] > 0),
        (df["TTM_영업CF"] / df["TTM_순이익"]) * 100, np.nan
    )

    # CAPEX 비율 (CAPEX / 영업CF × 100, 낮을수록 경자산 비즈니스)
    df["CAPEX비율(%)"] = np.where(
        pd.notna(df["TTM_CAPEX"]) & pd.notna(df["TTM_영업CF"]) & (df["TTM_영업CF"] > 0),
        (df["TTM_CAPEX"] / df["TTM_영업CF"]) * 100, np.nan
    )

    # 영업CF > 순이익 (이익 품질 플래그)
    df["이익품질_양호"] = np.where(
        pd.notna(df["TTM_영업CF"]) & pd.notna(df["TTM_순이익"]) & (df["TTM_순이익"] > 0),
        np.where(df["TTM_영업CF"] > df["TTM_순이익"], 1, 0), 0
    )

    # 부채상환능력 (영업CF / 부채총계, 높을수록 부채 상환 여력 큼)
    df["부채상환능력"] = np.where(
        pd.notna(df["TTM_영업CF"]) & pd.notna(df["부채"]) & (df["부채"] > 0),
        (df["TTM_영업CF"] / df["부채"]), np.nan
    )

    # F7: 주식 희석 없음 (발행주식수 미증가) — shares 데이터 활용
    if not shares_df.empty and "발행주식수" in shares_df.columns and "기준일" in shares_df.columns:
        for idx, row in df.iterrows():
            code = row["종목코드"]
            s = shares_df[shares_df["종목코드"] == code].sort_values("기준일")
            if len(s) >= 2:
                latest_shares = s["발행주식수"].iloc[-1]
                prev_shares = s["발행주식수"].iloc[-2]
                if pd.notna(latest_shares) and pd.notna(prev_shares) and prev_shares > 0:
                    df.at[idx, "F7_희석없음"] = 1 if latest_shares <= prev_shares else 0
        # F스코어 재계산 (F7 반영)
        if "F7_희석없음" in df.columns:
            df["F스코어"] = (
                df["F1_수익성"].fillna(0) + df["F2_영업CF"].fillna(0) +
                df["F3_ROA개선"].fillna(0) + df["F4_이익품질"].fillna(0) +
                df["F5_레버리지"].fillna(0) + df["F6_유동성"].fillna(0) +
                df["F7_희석없음"].fillna(0) + df["F8_매출총이익률"].fillna(0) +
                df["F9_자산회전율"].fillna(0)
            )

    # S-RIM
    Ke = 8.0
    df["적정주가_SRIM"] = np.where(
        (df["ROE(%)"] > Ke) & (df["BPS"] > 0),
        df["BPS"] + df["BPS"] * (df["ROE(%)"] - Ke) / Ke,
        np.where((df["BPS"] > 0), df["BPS"] * 0.9, np.nan)
    )
    df["괴리율(%)"] = ((df["적정주가_SRIM"] - df["종가"]) / df["종가"]) * 100

    # 검증 플래그
    df["PER_이상"] = np.where((df["PER"] < 0.5) | (df["PER"] > 500), "⚠️", "")

    # ── 스코어링 (NaN은 순위에서 제외 → NaN 유지, 스크리닝 단계에서 필터) ──
    df["S_PER"] = (1 - df["PER"].rank(pct=True, na_option='keep')) * 100
    df["S_PBR"] = (1 - df["PBR"].rank(pct=True, na_option='keep')) * 100
    df["S_ROE"] = df["ROE(%)"].rank(pct=True, na_option='keep') * 100

    df["S_매출CAGR"] = df["매출_CAGR"].rank(pct=True, na_option='keep') * 100
    df["S_영업이익CAGR"] = df["영업이익_CAGR"].rank(pct=True, na_option='keep') * 100
    df["S_순이익CAGR"] = df["순이익_CAGR"].rank(pct=True, na_option='keep') * 100

    # 연속성장: 각 항목 0~5년을 0~100으로 정규화 후 평균
    df["S_연속성장"] = (
        df["매출_연속성장"].fillna(0).clip(0, 5) / 5 * 100 +
        df["영업이익_연속성장"].fillna(0).clip(0, 5) / 5 * 100 +
        df["순이익_연속성장"].fillna(0).clip(0, 5) / 5 * 100
    ) / 3

    # 이익률 변동폭 연속값 사용 (이진 플래그 대신 실제 개선폭 반영)
    df["S_이익률개선"] = df["이익률_변동폭"].rank(pct=True, na_option='keep') * 100
    df["S_배당수익률"] = df["배당수익률(%)"].rank(pct=True, na_option='keep') * 100
    df["S_배당연속증가"] = df["배당_연속증가"].fillna(0).clip(0, 5) / 5 * 100
    df["S_괴리율"] = df["괴리율(%)"].rank(pct=True, na_option='keep') * 100
    df["S_F스코어"] = df["F스코어"].rank(pct=True, na_option='keep') * 100
    df["S_FCF수익률"] = df["FCF수익률(%)"].rank(pct=True, na_option='keep') * 100

    # 계절성 통제 스코어 (분기 YoY 기반)
    df["S_Q매출YoY"] = df["Q_매출_YoY(%)"].rank(pct=True, na_option='keep') * 100
    df["S_Q영업이익YoY"] = df["Q_영업이익_YoY(%)"].rank(pct=True, na_option='keep') * 100
    df["S_TTM매출YoY"] = df["TTM_매출_YoY(%)"].rank(pct=True, na_option='keep') * 100
    df["S_TTM영업이익YoY"] = df["TTM_영업이익_YoY(%)"].rank(pct=True, na_option='keep') * 100
    df["S_Q연속YoY"] = (
        df["Q_매출_연속YoY성장"].fillna(0).clip(0, 4) / 4 * 100 +
        df["Q_영업이익_연속YoY성장"].fillna(0).clip(0, 4) / 4 * 100
    ) / 2

    df["종합점수"] = (
        df["S_PER"].fillna(0) * 1.5 +
        df["S_PBR"].fillna(0) * 1.0 +
        df["S_ROE"].fillna(0) * 2.0 +
        df["S_매출CAGR"].fillna(0) * 2.0 +
        df["S_영업이익CAGR"].fillna(0) * 2.0 +
        df["S_순이익CAGR"].fillna(0) * 1.0 +
        df["S_연속성장"].fillna(0) * 1.0 +
        df["S_이익률개선"].fillna(0) * 1.0 +
        df["S_배당수익률"].fillna(0) * 0.3 +
        df["S_배당연속증가"].fillna(0) * 0.3 +
        df["S_괴리율"].fillna(0) * 1.0 +
        df["S_F스코어"].fillna(0) * 2.0 +
        df["S_FCF수익률"].fillna(0) * 1.5
    )

    return df


# ═════════════════════════════════════════════
# [v7] 기술적 지표 (주가 히스토리 기반)
# ═════════════════════════════════════════════

def calc_technical_indicators(df: pd.DataFrame, price_hist: pd.DataFrame) -> pd.DataFrame:
    """주가 히스토리로 기술적 지표를 계산하여 df에 병합.

    계산 지표:
      - 52주_최고대비(%): (종가 - 52주고가) / 52주고가 × 100 (음수 = 고점 대비 하락)
      - 52주_최저대비(%): (종가 - 52주저가) / 52주저가 × 100 (양수 = 저점 대비 상승)
      - MA20_이격도(%): (종가 / 20일이동평균 - 1) × 100
      - MA60_이격도(%): (종가 / 60일이동평균 - 1) × 100
      - RSI_14: 14일 RSI (50 이상 = 상승 모멘텀)
      - 거래대금_20일평균: 최근 20일 평균 거래대금
      - 거래대금_증감(%): (최근 5일 평균 / 20일 평균 - 1) × 100
      - 변동성_60일(%): 60일 수익률 표준편차 × √252 (연환산)
    """
    if price_hist.empty:
        log.warning("주가 히스토리 없음 — 기술적 지표 건너뜀")
        for col in ["52주_최고대비(%)", "52주_최저대비(%)", "MA20_이격도(%)", "MA60_이격도(%)",
                     "RSI_14", "거래대금_20일평균", "거래대금_증감(%)", "변동성_60일(%)"]:
            df[col] = np.nan
        return df

    tech_results = []
    for code in df["종목코드"].unique():
        ph = price_hist[price_hist["종목코드"] == code].copy()
        if ph.empty or len(ph) < 5:
            tech_results.append({"종목코드": code})
            continue

        ph = ph.sort_values("날짜")
        closes = ph["종가"].dropna()

        if len(closes) < 5:
            tech_results.append({"종목코드": code})
            continue

        latest_close = closes.iloc[-1]
        result = {"종목코드": code}

        # 52주 최고/최저 대비
        high_52w = ph["고가"].max() if "고가" in ph.columns else closes.max()
        low_52w = ph["저가"].min() if "저가" in ph.columns else closes.min()
        if pd.notna(high_52w) and high_52w > 0:
            result["52주_최고대비(%)"] = (latest_close - high_52w) / high_52w * 100
        if pd.notna(low_52w) and low_52w > 0:
            result["52주_최저대비(%)"] = (latest_close - low_52w) / low_52w * 100

        # 이동평균 이격도
        if len(closes) >= 20:
            ma20 = closes.iloc[-20:].mean()
            if ma20 > 0:
                result["MA20_이격도(%)"] = (latest_close / ma20 - 1) * 100
        if len(closes) >= 60:
            ma60 = closes.iloc[-60:].mean()
            if ma60 > 0:
                result["MA60_이격도(%)"] = (latest_close / ma60 - 1) * 100

        # RSI 14일
        if len(closes) >= 15:
            delta = closes.diff()
            gain = delta.where(delta > 0, 0.0)
            loss = (-delta.where(delta < 0, 0.0))
            avg_gain = gain.iloc[-14:].mean()
            avg_loss = loss.iloc[-14:].mean()
            if avg_loss > 0:
                rs = avg_gain / avg_loss
                result["RSI_14"] = 100 - (100 / (1 + rs))
            elif avg_gain > 0:
                result["RSI_14"] = 100.0
            else:
                result["RSI_14"] = 50.0

        # 거래대금 분석
        volumes = ph["거래량"].dropna() if "거래량" in ph.columns else pd.Series(dtype=float)
        amounts = ph["거래대금"].dropna() if "거래대금" in ph.columns else pd.Series(dtype=float)

        # 거래대금이 없으면 종가 × 거래량으로 추정
        if amounts.empty and not volumes.empty:
            amounts = (ph["종가"] * ph["거래량"]).dropna()

        if len(amounts) >= 20:
            avg_20d = amounts.iloc[-20:].mean()
            result["거래대금_20일평균"] = avg_20d
            if avg_20d > 0 and len(amounts) >= 5:
                avg_5d = amounts.iloc[-5:].mean()
                result["거래대금_증감(%)"] = (avg_5d / avg_20d - 1) * 100

        # 변동성 (60일, 연환산)
        if len(closes) >= 60:
            returns = closes.pct_change().dropna()
            if len(returns) >= 60:
                vol_60 = returns.iloc[-60:].std() * np.sqrt(252) * 100
                result["변동성_60일(%)"] = vol_60

        tech_results.append(result)

    tech_df = pd.DataFrame(tech_results)
    if tech_df.empty:
        return df

    # 병합
    tech_cols = [c for c in tech_df.columns if c != "종목코드"]
    for col in tech_cols:
        if col in df.columns:
            df = df.drop(columns=[col])

    df = df.merge(tech_df, on="종목코드", how="left")
    log.info("기술적 지표 계산 완료 (%d종목)", len(tech_df))
    return df


# ═════════════════════════════════════════════
# 스크리닝 (기존 2 + 신규 3 = 총 5개 필터)
# ═════════════════════════════════════════════

def apply_screen(df):
    """① 기본 우량주/저평가 스크리닝"""
    mask = (
        pd.notna(df["TTM_순이익"]) & (df["TTM_순이익"] > 0) &
        (df["ROE(%)"] >= 5) &
        (df["PER"].between(1, 50)) &
        (df["PBR"].between(0.1, 10)) &
        (df["매출_연속성장"] >= 2) &
        (df["순이익_연속성장"] >= 1) &
        (df["시가총액"] >= 50_000_000_000) &
        (df["PER_이상"] == "") &
        (df["F스코어"] >= 5)
    )
    return df[mask].sort_values("종합점수", ascending=False)


def apply_momentum_screen(df):
    """② 모멘텀/성장주 스크리닝 (계절성 통제 강화)"""
    mask = (
        pd.notna(df["매출_CAGR"]) &
        pd.notna(df["영업이익_CAGR"]) &
        ((df["매출_CAGR"] >= 15) | (df["영업이익_CAGR"] >= 15)) &
        (df["이익률_개선"] == 1) &
        (df["ROE(%)"] >= 5) &
        (df["TTM_순이익"] > 0) &
        (df["시가총액"] >= 50_000_000_000)
    )
    mom_df = df[mask].copy()
    if not mom_df.empty:
        mom_df["모멘텀_점수"] = (
            mom_df["매출_CAGR"].rank(pct=True) * 2.0 +
            mom_df["영업이익_CAGR"].rank(pct=True) * 2.5 +
            mom_df["ROE(%)"].rank(pct=True) * 1.5 +
            mom_df["영업이익률_최근"].rank(pct=True) * 1.0 +
            mom_df["이익률_개선"].rank(pct=True) * 0.5 +
            # 계절성 통제 지표 (분기 YoY)
            mom_df["Q_매출_YoY(%)"].fillna(0).rank(pct=True) * 2.0 +
            mom_df["Q_영업이익_YoY(%)"].fillna(0).rank(pct=True) * 2.0 +
            mom_df["Q_매출_연속YoY성장"].fillna(0).clip(0, 4).rank(pct=True) * 1.5 +
            mom_df["RSI_14"].fillna(50).rank(pct=True) * 1.0 +
            mom_df["MA20_이격도(%)"].fillna(0).rank(pct=True) * 1.0 +
            mom_df["거래대금_증감(%)"].fillna(0).rank(pct=True) * 0.5
        )
    if "모멘텀_점수" in mom_df.columns:
        return mom_df.sort_values("모멘텀_점수", ascending=False)
    return mom_df


def apply_garp_screen(df):
    """
    ③ GARP (Growth At Reasonable Price) — 피터 린치 스타일
    조건:
      - PEG < 1.5 (성장 대비 합리적 가격)
      - 매출 CAGR ≥ 10% (성장 확인)
      - ROE ≥ 12% (수익성 담보)
      - PER 5~30 (적자·극단 제외)
      - 시총 500억+ (소형주 제외)
    """
    mask = (
        pd.notna(df["PEG"]) & (df["PEG"] > 0) & (df["PEG"] < 1.5) &
        pd.notna(df["매출_CAGR"]) & (df["매출_CAGR"] >= 10) &
        pd.notna(df["ROE(%)"]) & (df["ROE(%)"] >= 12) &
        pd.notna(df["PER"]) & df["PER"].between(5, 30) &
        (df["시가총액"] >= 50_000_000_000) &
        (df["TTM_순이익"] > 0) &
        (df["PER_이상"] == "")
    )
    g = df[mask].copy()
    if not g.empty:
        g["GARP_점수"] = (
            (1 - g["PEG"].rank(pct=True)) * 3.0 +           # 낮은 PEG 선호
            g["매출_CAGR"].rank(pct=True) * 2.0 +            # 높은 매출 성장
            g["영업이익_CAGR"].rank(pct=True) * 1.5 +        # 높은 이익 성장
            g["ROE(%)"].rank(pct=True) * 2.0 +               # 높은 ROE
            (1 - g["PER"].rank(pct=True)) * 1.5 +            # 낮은 PER
            (1 - g["PBR"].clip(0.5, 10).rank(pct=True)) * 1.0 +  # 낮은 PBR
            g["현금전환율(%)"].fillna(100).clip(50, 200).rank(pct=True) * 1.0 +  # 현금 이익 품질
            g["F스코어"].fillna(0).rank(pct=True) * 0.5 +    # 재무건전성
            g["이익률_개선"].fillna(0) * 0.5 +               # 이익률 개선 보너스
            g["S_괴리율"].fillna(0) / 100 * 0.5              # S-RIM 저평가
        )
    if "GARP_점수" in g.columns:
        return g.sort_values("GARP_점수", ascending=False)
    return g


def apply_cashcow_screen(df):
    """
    ④ 캐시카우 (고수익 우량주) — 버핏 스타일
    조건 (안정적 지표 기반):
      - ROE ≥ 10% (높은 자본 수익성)
      - 영업이익률 ≥ 10% (높은 마진)
      - 부채비율 < 100% (또는 무차입)
      - 매출 연속성장 ≥ 1년
      - 시총 500억+
      - 흑자
      - 이익품질 양호 (영업CF > 순이익)
      - F스코어 ≥ 6 (재무 건전성)
    """
    mask = (
        pd.notna(df["ROE(%)"]) & (df["ROE(%)"] >= 10) &
        pd.notna(df["영업이익률(%)"]) & (df["영업이익률(%)"] >= 10) &
        (
            (pd.notna(df["부채비율(%)"]) & (df["부채비율(%)"] < 100)) |
            df["부채비율(%)"].isna()
        ) &
        (df["매출_연속성장"] >= 1) &
        (df["시가총액"] >= 50_000_000_000) &
        (df["TTM_순이익"] > 0) &
        (df["이익품질_양호"] == 1) &
        (df["F스코어"] >= 6)
    )
    c = df[mask].copy()
    if not c.empty:
        c["캐시카우_점수"] = (
            c["ROE(%)"].rank(pct=True) * 2.0 +                               # ROE
            c["영업이익률(%)"].rank(pct=True) * 2.0 +                         # 영업이익률
            (1 - c["부채비율(%)"].fillna(0).rank(pct=True)) * 1.5 +          # 저부채 선호
            c["FCF수익률(%)"].fillna(0).rank(pct=True) * 2.5 +               # FCF 수익률 (핵심)
            c["부채상환능력"].fillna(0).clip(0, 3).rank(pct=True) * 2.0 +    # 부채상환 여력
            c["매출_연속성장"].fillna(0).rank(pct=True) * 1.0 +              # 안정 성장
            (1 - c["PER"].clip(1, 100).rank(pct=True)) * 1.0 +              # 저PER
            c["배당수익률(%)"].rank(pct=True) * 0.5 +                         # 배당 보너스
            c["F스코어"].rank(pct=True) * 1.0 +                              # 재무건전성
            c["S_괴리율"].fillna(0) / 100 * 0.5                              # S-RIM 저평가
        )
    if "캐시카우_점수" in c.columns:
        return c.sort_values("캐시카우_점수", ascending=False)
    return c


def apply_turnaround_screen(df):
    """
    ⑤ 턴어라운드 (실적 반등) — 역발상 투자
    조건 (OR):
      A) 적자→흑자 전환 (흑자전환 == 1)
      B) 영업이익률 +5%p 이상 급개선 (이익률_급개선 == 1)
    공통:
      - 현재 순이익 > 0 (현재 흑자)
      - 시총 300억+ (소형주 포함 — 턴어라운드는 초기 발굴)
    """
    mask = (
        (
            (df["흑자전환"] == 1) |
            (df["이익률_급개선"] == 1)
        ) &
        (df["TTM_순이익"] > 0) &
        (df["시가총액"] >= 30_000_000_000)
    )
    t = df[mask].copy()
    if not t.empty:
        t["턴어라운드_점수"] = (
            t["이익률_변동폭"].fillna(0).rank(pct=True) * 2.0 +       # 이익률 개선폭
            t["매출_CAGR"].rank(pct=True) * 2.0 +                     # 매출 성장 (더 중요)
            t["ROE(%)"].rank(pct=True) * 1.5 +                        # ROE
            t["흑자전환"].fillna(0) * 2.0 +                           # 흑전 보너스
            (1 - t["PER"].clip(0, 100).rank(pct=True)) * 1.0 +       # 저PER
            t["이익률_급개선"].fillna(0) * 1.5 +                      # 급개선 보너스
            (1 - t["RSI_14"].fillna(50).rank(pct=True)) * 1.0 +      # 과매도 선호
            (1 - t["52주_최고대비(%)"].fillna(0).abs().rank(pct=True)) * 1.0 +  # 저점 매수 기회
            t["F스코어"].fillna(0).rank(pct=True) * 0.5 +             # 최소 재무건전성
            t["S_괴리율"].fillna(0) / 100 * 0.5                      # S-RIM 저평가
        )
    if "턴어라운드_점수" in t.columns:
        return t.sort_values("턴어라운드_점수", ascending=False)
    return t


def apply_dividend_growth_screen(df):
    """
    ⑥ 배당 성장주 (배당금 & 수익 지속 증가) — 배당 성향 관리 기업
    조건:
      - 순이익 연속성장 ≥ 2년
      - 배당금(DPS) 연속증가 ≥ 1년
      - DPS_CAGR > 0 (배당금 성장)
      - ROE ≥ 5% (수익성)
      - 배당수익률 > 0 (배당 중시 기업)
      - 시총 300억+
      - 현재 흑자
      - 수익 + 배당 동반증가 확인 (배당_수익동반증가 == 1)
    """
    mask = (
        (df["순이익_연속성장"] >= 2) &
        (df["배당_연속증가"] >= 1) &
        pd.notna(df["DPS_CAGR"]) & (df["DPS_CAGR"] > 0) &
        pd.notna(df["ROE(%)"]) & (df["ROE(%)"] >= 5) &
        (df["배당수익률(%)"] > 0) &
        (df["시가총액"] >= 30_000_000_000) &
        (df["TTM_순이익"] > 0) &
        (df["배당_수익동반증가"] == 1)
    )
    d = df[mask].copy()
    if not d.empty:
        d["배당성장_점수"] = (
            d["DPS_CAGR"].rank(pct=True) * 3.0 +                    # 배당 성장률 (핵심)
            d["순이익_CAGR"].rank(pct=True) * 2.5 +                   # 수익 성장률
            d["배당_연속증가"].fillna(0).rank(pct=True) * 2.0 +        # 연속 배당 증가
            d["순이익_연속성장"].fillna(0).rank(pct=True) * 2.0 +      # 연속 수익 증가
            d["ROE(%)"].rank(pct=True) * 1.5 +                        # 자본 수익성
            d["배당수익률(%)"].rank(pct=True) * 1.5 +                  # 배당 수익률
            (1 - d["부채비율(%)"].fillna(0).rank(pct=True)) * 1.0 +   # 저부채 선호
            d["F스코어"].fillna(0).rank(pct=True) * 0.5 +             # 재무건전성
            (1 - d["PER"].clip(1, 100).rank(pct=True)) * 0.5         # 저PER
        )
    if "배당성장_점수" in d.columns:
        return d.sort_values("배당성장_점수", ascending=False)
    return d


# ═════════════════════════════════════════════
# 엑셀 저장
# ═════════════════════════════════════════════

def save_to_excel(df, filepath, sheet_name="Result"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    col_groups = {
        "기본정보": ["종목코드", "종목명", "종가", "시가총액", "상장주식수"],
        "주요지표": ["PER", "PBR", "PSR", "PEG", "ROE(%)", "EPS", "BPS",
                    "부채비율(%)", "영업이익률(%)", "이익수익률(%)", "FCF수익률(%)",
                    "배당수익률(%)", "이익품질_양호"],
        "배당": ["DPS_최근", "DPS_CAGR", "배당_연속증가", "배당_수익동반증가"],
        "점수": ["종합점수", "모멘텀_점수", "GARP_점수", "캐시카우_점수", "턴어라운드_점수", "배당성장_점수"],
        "성장추세(연간)": ["매출_CAGR", "영업이익_CAGR", "순이익_CAGR",
                        "매출_연속성장", "영업이익_연속성장", "순이익_연속성장",
                        "이익률_개선", "이익률_급개선", "이익률_변동폭",
                        "흑자전환", "영업이익률_최근", "영업이익률_전년"],
        "성장추세(분기YoY)": ["최근분기",
                            "Q_매출_YoY(%)", "Q_영업이익_YoY(%)", "Q_순이익_YoY(%)",
                            "Q_매출_연속YoY성장", "Q_영업이익_연속YoY성장", "Q_순이익_연속YoY성장",
                            "TTM_매출_YoY(%)", "TTM_영업이익_YoY(%)", "TTM_순이익_YoY(%)"],
        "밸류에이션": ["적정주가_SRIM", "괴리율(%)"],
        "TTM_원본": ["TTM_매출", "TTM_영업이익", "TTM_순이익", "TTM_영업CF", "자본", "부채"],
    }

    ordered_cols = []
    for g in col_groups.values():
        for c in g:
            if c in df.columns:
                ordered_cols.append(c)
    for c in df.columns:
        if c not in ordered_cols and not c.startswith("S_"):
            ordered_cols.append(c)

    export_df = df[ordered_cols].copy()

    fills = {
        "기본정보": PatternFill("solid", fgColor="D6E4F0"),
        "주요지표": PatternFill("solid", fgColor="E2EFDA"),
        "점수": PatternFill("solid", fgColor="C6EFCE"),
        "성장추세(연간)": PatternFill("solid", fgColor="FFF2CC"),
        "성장추세(분기YoY)": PatternFill("solid", fgColor="FCE4D6"),
        "밸류에이션": PatternFill("solid", fgColor="DAEEF3"),
        "TTM_원본": PatternFill("solid", fgColor="F2DCDB"),
    }

    header_font = Font(name="맑은 고딕", bold=True, size=10)
    data_font = Font(name="맑은 고딕", size=9)
    thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))

    for col_idx, col_name in enumerate(ordered_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        for grp, cols in col_groups.items():
            if col_name in cols:
                cell.fill = fills[grp]
                break

    for row_idx, (_, row_data) in enumerate(export_df.iterrows(), 2):
        for col_idx, col_name in enumerate(ordered_cols, 1):
            val = row_data[col_name]
            if pd.isna(val): val = None
            elif isinstance(val, (np.floating, float)): val = round(float(val), 2)
            elif isinstance(val, (np.integer,)): val = int(val)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border

            if col_name in ["시가총액", "종가", "EPS", "BPS", "적정주가_SRIM"]:
                cell.number_format = '#,##0'
            elif "%" in col_name or "CAGR" in col_name:
                cell.number_format = '#,##0.00'
            elif "점수" in col_name:
                cell.number_format = '#,##0.0'

    for col_idx, col_name in enumerate(ordered_cols, 1):
        width = 12
        if col_name == "종목명": width = 18
        elif "점수" in col_name: width = 14
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "C2"
    wb.save(filepath)
    log.info(f"💾 저장: {filepath}")


# ═════════════════════════════════════════════
# 메인 실행
# ═════════════════════════════════════════════

def run():
    master = load_table("master")
    daily = load_table("daily")
    fs = load_table("financial_statements")
    ind = load_table("indicators")
    shares = load_table("shares")

    if daily.empty:
        log.error("❌ daily 없음")
        return

    ind = preprocess_indicators(ind)
    multiplier = detect_unit_multiplier(ind)
    anal_df = analyze_all(fs, ind)

    full_df = calc_valuation(daily, anal_df, multiplier, shares)

    # ── 6개 결과물 저장 ──
    # 1. 전체
    save_to_excel(full_df.sort_values("종합점수", ascending=False),
                  DATA_DIR / "quant_all_stocks.xlsx", "전체종목")

    # 2. 우량주
    screened = apply_screen(full_df)
    save_to_excel(screened, DATA_DIR / "quant_screened.xlsx", "우량주")

    # 3. 모멘텀
    momentum_df = apply_momentum_screen(full_df)
    save_to_excel(momentum_df, DATA_DIR / "quant_momentum.xlsx", "모멘텀")

    # 4. GARP (성장+합리적 가격)
    garp_df = apply_garp_screen(full_df)
    save_to_excel(garp_df, DATA_DIR / "quant_GARP.xlsx", "GARP")

    # 5. 캐시카우 (현금흐름 우량)
    cashcow_df = apply_cashcow_screen(full_df)
    save_to_excel(cashcow_df, DATA_DIR / "quant_cashcow.xlsx", "캐시카우")

    # 6. 턴어라운드 (실적 반등)
    turnaround_df = apply_turnaround_screen(full_df)
    save_to_excel(turnaround_df, DATA_DIR / "quant_turnaround.xlsx", "턴어라운드")

    # 7. 배당 성장 (수익 + 배당 동반증가)
    dividend_growth_df = apply_dividend_growth_screen(full_df)
    save_to_excel(dividend_growth_df, DATA_DIR / "quant_dividend_growth.xlsx", "배당성장")

    # ── 요약 출력 ──
    print("\n" + "=" * 80)
    print(f"📏 단위 보정: {multiplier:,.0f}")
    print(f"📊 분석 종목:             {len(full_df):,}개")
    print(f"✅ 우량주 스크리닝:        {len(screened):,}개")
    print(f"🚀 모멘텀 (고성장):        {len(momentum_df):,}개")
    print(f"📈 GARP (성장+가치):       {len(garp_df):,}개")
    print(f"💵 캐시카우 (현금흐름):    {len(cashcow_df):,}개")
    print(f"🔄 턴어라운드 (반등):      {len(turnaround_df):,}개")
    print(f"💰 배당 성장 (수익+배당):  {len(dividend_growth_df):,}개")
    print("=" * 80)

    # TOP 10 출력
    if len(momentum_df) > 0:
        print("\n🚀 모멘텀 TOP 10:")
        cols = ["종목명", "매출_CAGR", "영업이익_CAGR", "영업이익률_최근", "ROE(%)", "모멘텀_점수"]
        valid = [c for c in cols if c in momentum_df.columns]
        print(momentum_df[valid].head(10).to_string(index=False, float_format="%.1f"))

    if len(garp_df) > 0:
        print("\n📈 GARP TOP 10:")
        cols = ["종목명", "PEG", "매출_CAGR", "ROE(%)", "PER", "GARP_점수"]
        valid = [c for c in cols if c in garp_df.columns]
        print(garp_df[valid].head(10).to_string(index=False, float_format="%.1f"))

    if len(cashcow_df) > 0:
        print("\n💵 캐시카우 TOP 10:")
        cols = ["종목명", "FCF수익률(%)", "영업이익률(%)", "부채비율(%)", "ROE(%)", "캐시카우_점수"]
        valid = [c for c in cols if c in cashcow_df.columns]
        print(cashcow_df[valid].head(10).to_string(index=False, float_format="%.1f"))

    if len(turnaround_df) > 0:
        print("\n🔄 턴어라운드 TOP 10:")
        cols = ["종목명", "흑자전환", "이익률_급개선", "이익률_변동폭", "영업이익률_최근", "ROE(%)", "턴어라운드_점수"]
        valid = [c for c in cols if c in turnaround_df.columns]
        print(turnaround_df[valid].head(10).to_string(index=False, float_format="%.1f"))

    if len(dividend_growth_df) > 0:
        print("\n💰 배당 성장 TOP 10:")
        cols = ["종목명", "DPS_CAGR", "순이익_CAGR", "배당수익률(%)", "배당_연속증가", "ROE(%)", "배당성장_점수"]
        valid = [c for c in cols if c in dividend_growth_df.columns]
        print(dividend_growth_df[valid].head(10).to_string(index=False, float_format="%.1f"))


if __name__ == "__main__":
    run()